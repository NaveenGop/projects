import sys, os, re
import openpyxl


def take_range_input(user_input, s):
	if user_input.lower() in ['q', 'x', 'exit']:
		print "USER QUIT"
		sys.exit()
	elif user_input.startswith('0'):
		item_list = [x for x in range(1, s + 1)]
	else:
		# FORMAT THE INDICATED RANGES
		item_list = []
		pattern = re.compile("[0-9]+\\s*(-\\s*\\d+)|[0-9]+(?=\\s*,)|\\d+(?!.)")
		it = pattern.search(user_input)
		while it is not None:
			item_list.append(user_input[it.start():it.end()].replace(" ", ""))
			user_input = user_input[it.end() + 1:]
			it = pattern.search(user_input)
		temp = set()
		for x in item_list:
			if '-' in x:
				a = x.split('-')
				temp.update([y for y in range(int(a[0]), int(a[1]) + 1)])
			else:
				temp.add(int(x))
		item_list = [x for x in sorted(temp) if x <= s]
		del temp
	print item_list
	return item_list


def main(excel=r'coo3.xlsx'):
	CONFLUENCE_DIR = r'\\amr\ec\proj\fm\NSG\SE\Cliffdale_Refresh_VE\SI\Confluence_Results' # r'D:\CofluentResults'
	workbook = openpyxl.Workbook()
	del workbook['Sheet']
	folder_list = []
	file_list = []
	sheet_list = [u'IOpsResults', u'MiBps_Results', u'WAF_Output', u'Cofluent Workloads', u'Cofluent Models', u'Perf Per Watt', u'Sim Stats']
	user_sheets = []

	COMMAND_LIST = []
	VIEW_DIR = ['vd', 'viewdirectory']
	COMMAND_LIST.append(VIEW_DIR)
	CHANGE_DIR = ['cd', 'changedirectory']
	COMMAND_LIST.append(CHANGE_DIR)
	SELECT_FOLDERS = ['sf', 'selectfolders']
	COMMAND_LIST.append(SELECT_FOLDERS)
	FIND_WORKBOOKS = ['fw', 'findworkbook']
	COMMAND_LIST.append(FIND_WORKBOOKS)
	SELECT_SHEETS = ['ss', 'selectsheet']
	RUN_PARSE = ['rp', 'runparse']
	COMMAND_LIST.append(RUN_PARSE)
	COMMAND_LIST.append(SELECT_SHEETS)
	EXIT_COMMANDS = ['q', 'x', 'exit']
	COMMAND_LIST.append(EXIT_COMMANDS)

	while True:
		user_input = raw_input("\nEnter command ('help' to list all commands): ")
		user_input = user_input.lower().strip()

		if user_input in ['h', 'help']:
			print("{}\nALL commands are NOT case sensitive\n{}".format('*'*34, '*'*34))
			print("\nCOMMAND LIST")
			for x in COMMAND_LIST:
				print x
			continue
		elif user_input in VIEW_DIR:
			# CHECK DIRECTORY OF CONFLUENCE RESULTS
			print "Confluent Results Directory: {}".format(CONFLUENCE_DIR)
			continue
		elif user_input in CHANGE_DIR:
			# noinspection PyShadowingBuiltins
			input = raw_input("Enter new Confluent Results Directory: ")
			if input in EXIT_COMMANDS:
				continue
			else:
				# LIST THE FOLDERS IN CONFLUENCE RESULTS DIRECTORY
				if os.path.isdir(input):
					CONFLUENCE_DIR = input
				else:
					print 'INVALID PATH NAME. TRY AGAIN.'
				continue
		elif user_input in SELECT_FOLDERS:
			print "{}\n{}\n{}".format('*' * 45, 'RESULT FOLDERS'.center(45, ' '), '*' * 45)
			s = [x for x in os.listdir(CONFLUENCE_DIR) if os.path.isdir(CONFLUENCE_DIR+'\\'+x)]
			for x in range(1, len(s) + 1):
				y = s[x - 1].split('_', 2)
				try:
					print "{}. {} (Time={} at {})".format(x, y[-1].replace('_', ' '), y[0], y[1])
				except IndexError:
					print "{}. {} (Irregular Folder Name)".format(x, s[x-1])
			# noinspection PyShadowingBuiltins
			input = raw_input("\nEnter range of numbers with \'-\' or a list of index numbers and ranges separated by a \',\'. Enter 0 to add all.\n\t")
			folder_list = take_range_input(input, len(s))
			continue
		elif user_input in FIND_WORKBOOKS:
			# SCANS FOR RESULTS.XLSX FOR INDICATED RANGES
			s = [x for x in os.listdir(CONFLUENCE_DIR) if os.path.isdir(CONFLUENCE_DIR+'\\'+x)]
			file_list = []
			for index in folder_list:
				path = CONFLUENCE_DIR + '\\' + s[index - 1]
				c = s[index - 1].split("_", 2)
				for b in os.listdir(path):
					if b.startswith(c[2] + "_Results") and b.endswith(c[1] + ".xlsx"):
						file_list.append(path + '\\' + b)
			if len(file_list) == 0:
				print "No results found."
			print file_list
			continue
		elif user_input in SELECT_SHEETS:
			# INDICATE AND SELECT SHEETS IN WORKBOOK
			for x in range(1, len(sheet_list) + 1):
				print "{}. {}".format(x, sheet_list[x - 1])
			# noinspection PyShadowingBuiltins
			input = raw_input("\nEnter range of numbers with \'-\' or a list of index numbers and ranges separated by a \',\'. Enter 0 to add all.\n\t")
			user_sheets = take_range_input(input, len(sheet_list))
			continue
		elif user_input in RUN_PARSE:
			# CHECK THAT WORKBOOKS HAVE BEEN SELECTED
			if len(file_list) == 0:
				print 'NO WORKBOOKS.'
				continue
			# CHECK THAT SHEETS HAVE BEEN SELECTED
			elif len(user_sheets) == 0:
				print 'NO SELECTED SHEETS. ADD SHEETS.'
				continue

			# CREATE DICTIONARY FOR CURRENT COLUMN
			current_col = {}
			for sheet in user_sheets:
				sleet = sheet_list[sheet-1]
				workbook.create_sheet(sleet)
				current_col[sleet] = 1

			second_col = {}
			for sheet in [u'IOpsResults', u'MiBps_Results']:
				second_col[sheet] = 1
			# CYCLE THROUGH WORKBOOKS
			for file_path in file_list:
				wb = openpyxl.load_workbook(file_path, data_only=True)
				# CYCLE THROUGH WORKSHEETS
				for sheet in user_sheets:
					sleet = sheet_list[sheet-1]
					try:
						ws = wb[sleet]
					except KeyError:
						print "{} does not have the sheet \'{}\'".format(file_path.rsplit('\\', 1)[-1], sleet)
						continue
					myws = workbook[sleet]
					# MAKE TITLE CELL
					myws.cell(row=1, column=current_col[sleet], value=file_path.rsplit('\\', 1)[-1].rstrip('.xlsx').replace('_', ' '))
					myws.merge_cells(start_row=1, start_column=current_col[sleet], end_row=1, end_column=current_col[sleet]+ws.max_column-1)
					# ITERATE THROUGH ROWS
					for row in ws.iter_rows(min_col=1, min_row=1, max_col=ws.max_column, max_row=ws.max_row):
						for cell in row:
							myws.cell(row=cell.row+1, column=current_col[sleet]+cell.col_idx-1, value=cell.value)
					# myws.cell(row=cell.row + 2, column=current_col[sleet] + cell.col_idx - 1, value=None)
					if sleet == u'IOpsResults' or sleet == u'MiBps_Results':
						r = myws.max_row + 2 if file_path == file_list[0] else myws.max_row+1
						for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=2, max_row=ws.max_row):
							if any(x.value is None for x in col):
								break
							if all(x.value == u'NULL' for x in col):
								continue
							for cell in col:
								if type(cell.value) == float or type(cell.value) == int or type(cell.value) == long:
									myws.cell(row=r, column=1, value=ws.cell(row=cell.row, column=1).value)
									myws.cell(row=r, column=1 + 1, value=myws.cell(row=1, column=current_col[sleet]).value)
									myws.cell(row=r, column=1 + 2, value=ws.cell(row=1, column=cell.col_idx).value)
									myws.cell(row=r, column=1 + 3, value=cell.value)
									r += 1
						second_col[sleet] += 5

					if ws.max_column > 4:
						current_col[sleet] += ws.max_column+1
					else:
						current_col[sleet] += 5
			workbook.save(excel)
			continue
		elif user_input in EXIT_COMMANDS:
			print "Program Exiting......."
			break
		else:
			print "INVALID INPUT. TRY AGAIN."
			continue

if __name__ == '__main__':
	if len(sys.argv) == 2:
		main(sys.argv[1])
	else:
		main()
