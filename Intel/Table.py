import sys, os, re
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
	"""
	Apply styles to a range of cells as if they were a single cell.

	:param alignment:
	:param ws:  Excel worksheet instance
	:param cell_range: An excel range to style (e.g. A1:F20)
	:param border: An openpyxl Border
	:param fill: An openpyxl PatternFill or GradientFill
	:param font: An openpyxl Font object
	"""

	top = Border(top=border.top)
	left = Border(left=border.left)
	right = Border(right=border.right)
	bottom = Border(bottom=border.bottom)

	first_cell = ws[cell_range.split(":")[0]]
	if alignment:
		first_cell.alignment = alignment

	rows = ws[cell_range]
	if font:
		first_cell.font = font

	for cell in rows[0]:
		cell.border = cell.border + top
	for cell in rows[-1]:
		cell.border = cell.border + bottom

	for row in rows:
		l = row[0]
		r = row[-1]
		l.border = l.border + left
		r.border = r.border + right
		if fill:
			for c in row:
				c.fill = fill


def take_range_input(user_input, s):
	if user_input.lower() in ['q', 'x', 'exit']:
		print "USER QUIT"
		sys.exit()
	elif user_input.startswith('0'):
		item_list = [x for x in range(1, s + 1)]
	else:
		# FORMAT THE INDICATED RANGES
		item_list = []
		pattern = re.compile("[0-9]+\\s*(-\\s*\\d+)|[0-9]+(?=\\s*,)|\\d+(?!.)")
		it = pattern.search(user_input)
		while it is not None:
			item_list.append(user_input[it.start():it.end()].replace(" ", ""))
			user_input = user_input[it.end() + 1:]
			it = pattern.search(user_input)
		temp = set()
		for x in item_list:
			if '-' in x:
				a = x.split('-')
				temp.update([y for y in range(int(a[0]), int(a[1]) + 1)])
			else:
				temp.add(int(x))
		item_list = [x for x in sorted(temp) if x <= s]
		del temp
	print item_list
	return item_list


def assign_style(cell, border=None, alignment=None, fill=None, font=None):
	if border is not None:
		cell.border = border
	if alignment is not None:
		cell.alignment = alignment
	if fill is not None:
		cell.fill = fill
	if font is not None:
		cell.font = font


def main(excel='update.xlsx'):
	CONFLUENCE_DIR = r'D:\CofluentResults'  # r'\\amr\ec\proj\fm\NSG\SE\Cliffdale_Refresh_VE\SI\Confluence_Results'
	folder_list = []
	file_list = []
	workbook = Workbook()
	del workbook['Sheet']
	sheet_list = [u'IOpsResults', u'MiBps_Results', u'Perf Per Watt']

	COMMAND_LIST = []
	VIEW_DIR = ['vd', 'viewdirectory']
	COMMAND_LIST.append(VIEW_DIR)
	CHANGE_DIR = ['cd', 'changedirectory']
	COMMAND_LIST.append(CHANGE_DIR)
	SELECT_FOLDERS = ['sf', 'selectfolders']
	COMMAND_LIST.append(SELECT_FOLDERS)
	FIND_WORKBOOKS = ['fw', 'findworkbook']
	COMMAND_LIST.append(FIND_WORKBOOKS)
	RUN_PARSE = ['rp', 'runparse']
	COMMAND_LIST.append(RUN_PARSE)
	EXIT_COMMANDS = ['q', 'x', 'exit']
	COMMAND_LIST.append(EXIT_COMMANDS)
	while True:
		user_input = raw_input("\nEnter command ('help' to list all commands): ")
		user_input = user_input.lower().strip()

		if user_input in ['h', 'help']:
			print("{}\nALL commands are NOT case sensitive\n{}".format('*'*34, '*'*34))
			print("\nCOMMAND LIST")
			for x in COMMAND_LIST:
				print x
			continue
		elif user_input in VIEW_DIR:
			# CHECK DIRECTORY OF CONFLUENCE RESULTS
			print "Confluent Results Directory: {}".format(CONFLUENCE_DIR)
			continue
		elif user_input in CHANGE_DIR:
			# noinspection PyShadowingBuiltins
			input = raw_input("Enter new Confluent Results Directory: ")
			if input in EXIT_COMMANDS:
				continue
			else:
				# LIST THE FOLDERS IN CONFLUENCE RESULTS DIRECTORY
				if os.path.isdir(input):
					CONFLUENCE_DIR = input
				else:
					print 'INVALID PATH NAME. TRY AGAIN.'
				continue
		elif user_input in SELECT_FOLDERS:
			print "{}\n{}\n{}".format('*' * 45, 'RESULT FOLDERS'.center(45, ' '), '*' * 45)
			s = [x for x in os.listdir(CONFLUENCE_DIR) if os.path.isdir(CONFLUENCE_DIR+'\\'+x)]
			for x in range(1, len(s) + 1):
				y = s[x - 1].split('_', 2)
				try:
					print "{}. {} (Time={} at {})".format(x, y[-1].replace('_', ' '), y[0], y[1])
				except IndexError:
					print "{}. {} (Irregular Folder Name)".format(x, s[x-1])
			# noinspection PyShadowingBuiltins
			input = raw_input("\nEnter range of numbers with \'-\' or a list of index numbers and ranges separated by a \',\'. Enter 0 to add all.\n\t")
			folder_list = take_range_input(input, len(s))
			continue
		elif user_input in FIND_WORKBOOKS:
			# SCANS FOR RESULTS.XLSX FOR INDICATED RANGES
			s = [x for x in os.listdir(CONFLUENCE_DIR) if os.path.isdir(CONFLUENCE_DIR+'\\'+x)]
			file_list = []
			for index in folder_list:
				path = CONFLUENCE_DIR + '\\' + s[index - 1]
				c = s[index - 1].split("_", 2)
				for b in os.listdir(path):
					if b.startswith(c[2] + "_Results") and b.endswith(c[1] + ".xlsx"):
						file_list.append(path + '\\' + b)
			if len(file_list) == 0:
				print "No results found."
			print file_list
			continue
		elif user_input in RUN_PARSE:
			# CHECK THAT WORKBOOKS HAVE BEEN SELECTED
			if len(file_list) == 0:
				print 'NO WORKBOOKS.'
				continue
			di = []
			for file_path in file_list:
				name = file_path.rsplit('\\', 1)[-1].rsplit('_', 2)[0].replace('_', ' ')
				workbook.create_sheet(name if len(name) < 31 else name[:31])
				wb = load_workbook(file_path, data_only=True)
				ws = wb[u'IOpsResults']
				test_row = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column+1) if ws.cell(row=1, column=col).value is not None]
				test_names = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row+1) if ws.cell(row=row, column=1).value is not None]
				row_list = []
				for x in test_names:
					row_dic = {'test': x}
					a = re.match(u'\\d{0,3}k|\\d{0,3}B', x)
					row_dic['size'] = a.group(0)
					row_dic['rw'] = x[a.end(): x.find(u'Qd')]
					row_list.append(row_dic)

				# CYCLE THROUGH WORKSHEETS
				for sheet in sheet_list:
					if sheet in [u'IOpsResults', u'MiBps_Results']:
						try:
							ws = wb[sheet]
						except KeyError:
							print "{} does not have the sheet \'{}\'".format(file_path.rsplit('\\', 1)[-1], sheet)
							continue
						for col in ws.iter_cols(min_col=2, min_row=2, max_col=ws.max_column, max_row=ws.max_row):
							if any(x.value is None for x in col):
								break
							if sheet == sheet_list[0]:
								if all(x.value == u'NULL' for x in col):
									test_row.remove(ws.cell(row=1, column=col[0].col_idx).value)
									continue
							for cell in col:
								if cell.value is not None and type(cell.value) != unicode:
									a = int(cell.value) if type(cell.value) == float else cell.value
									temp_dict = None
									if (sheet == u'IOpsResults' and u'Rand' in ws.cell(row=cell.row, column=1).value) \
										or (sheet == u'MiBps_Results' and u'Seq' in ws.cell(row=cell.row, column=1).value):
											temp_dict = {'sheet': sheet,
													'name': ws.cell(row=1, column=cell.col_idx).value,
													'test': ws.cell(row=cell.row, column=1).value
													, 'value': a}
									if temp_dict is not None:
										di.append(temp_dict)
					elif sheet in [u'Perf Per Watt']:
						try:
							ws = wb[sheet]
						except KeyError:
							print "{} does not have the sheet \'{}\'".format(file_path.rsplit('\\', 1)[-1], sheet)
							continue
						for row in ws.iter_rows(min_col=2, min_row=2, max_col=ws.max_column, max_row=ws.max_row):
							if all(x.value is None for x in row):
								break
							for cell in row:
								if cell.value is not None and type(cell.value) != unicode:
									a = int(cell.value) if type(cell.value) == float else cell.value
									temp_dict = {'sheet': sheet, 'name': ws.cell(row=1, column=cell.col_idx).value, 'test': ws.cell(row=cell.row, column=1).value
												, 'value': a}
									di.append(temp_dict)


				myws = workbook[name if len(name) < 31 else name[:31]]
				arial_font = Font(name='Arial', size=8, color='333333')
				bold_font = Font(name='Arial', size=8, color='333333', bold=True)
				lo = Side(style='double', color='aeaaaa')
				border = Border(bottom=lo, top=lo, left=lo, right=lo)
				alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
				curr_col, temp, row = 0, 0, 0
				for table_iter in range(0, 2):
					if len([x['test'] for x in row_list if 'Rand' in x['test']]) > 0:
						assign_style(myws.cell(row=2, column=curr_col+1, value='Transfer Size'), alignment=alignment, border=border, font=bold_font)
						assign_style(myws.cell(row=2, column=curr_col+2, value='RANDOM Work Loads'), alignment=alignment, border=border, font=bold_font)
						assign_style(myws.cell(row=2, column=curr_col+3, value='QD'), alignment=alignment, border=border, font=bold_font)
						for x in range(0, 2*len(test_row)):
							if x < len(test_row):
								myws.cell(row=2, column=curr_col+4+x, value=test_row[x].replace('_', ' ') + u' (IOPS)')
								a = curr_col+4+x
							else:
								myws.cell(row=2, column=curr_col+4+len(test_row)+x % len(test_row), value=test_row[x-len(test_row)].replace('_', ' ') + u' (mW)')
								a = curr_col+4+len(test_row)+x % len(test_row)
							assign_style(myws.cell(row=2, column=a), font=bold_font, alignment=alignment, border=border, fill=PatternFill(start_color='e0f0ff', end_color='e0f0ff', patternType='solid'))
					size_row = 3
					temp_size = [x['size'] for x in row_list if u'Rand' in x['rw']]
					size = sorted(set(temp_size), key=lambda x: (x[-1].lower(), len(x), int(x[0])))
					for y in size:
						myws.cell(row=size_row, column=curr_col+1, value=y.upper())
						myws.merge_cells(start_row=size_row, start_column=curr_col+1, end_row=size_row+temp_size.count(y)-1, end_column=curr_col+1)
						if temp_size.count(y) != 1:
							style_range(myws, '{0}{1}:{0}{2}'.format(hex(65+curr_col)[2:].decode('hex'), size_row, size_row+temp_size.count(y)-1), font=bold_font, alignment=alignment, border=border)
						else:
							assign_style(myws.cell(row=size_row, column=curr_col+1), font=bold_font, alignment=alignment, border=border)
						rw_size = size_row

						temp_rw = [x['rw'] for x in row_list if u'Rand' in x['rw'] and y == x['size']]
						rw = sorted(set(temp_rw), key=lambda x: ('70' in x, 'Read' in x, 'Write' in x))
						for z in rw:
							if z == 'RandRead':
								myws.cell(row=rw_size, column=curr_col+2, value='Random Read')
							elif z == 'RandWrite':
								myws.cell(row=rw_size, column=curr_col+2, value='Random Write')
							else:
								myws.cell(row=rw_size, column=curr_col+2, value='70% Read 30% Write')
							myws.merge_cells(start_row=rw_size, start_column=curr_col+2, end_row=rw_size + temp_rw.count(z) - 1, end_column=curr_col+2)
							if temp_rw.count(z) != 1:
								style_range(myws, '{0}{1}:{0}{2}'.format(
									hex(66+curr_col)[2:].decode('hex'), rw_size,
									rw_size + temp_rw.count(z) - 1), font=arial_font, alignment=alignment, border=border)
							else:
								assign_style(myws.cell(row=rw_size, column=curr_col+2), font=arial_font, border=border, alignment=alignment)
							qd_row = rw_size
							temp_qd = [x['test'][x['test'].find(u'Qd'):x['test'].find(u'_')] for x in row_list if u'Rand' in x['rw'] and y == x['size'] and z == x['rw']]
							qd = sorted(set(temp_qd), key=lambda x: (len(x), x))
							for e in qd:
								myws.cell(row=qd_row, column=curr_col+3, value=e.upper())
								myws.merge_cells(start_row=qd_row, start_column=curr_col+3, end_row=qd_row + temp_qd.count(e) - 1, end_column=curr_col+3)
								if temp_qd.count(e) != 1:
									style_range(myws, '{0}{1}:{0}{2}'.format(
										hex(67+curr_col)[2:].decode('hex'), qd_row,
										qd_row + temp_qd.count(e) - 1), font=arial_font, alignment=alignment, border=border)
								else:
									assign_style(myws.cell(row=qd_row, column=curr_col+3), font=arial_font, border=border, alignment=alignment)

								a = [x['test'] for x in row_list if x['test'].startswith(y + z + e + '_')]
								for l in range(len(test_row)):
									for x in di:
										if x['sheet'] == u'IOpsResults' and x['name'] == test_row[l]:
											j = curr_col+4+l
											for c in a:
												if x['test'] == c:
													if not table_iter:
														myws.cell(row=qd_row, column=j, value=x['value'])
													else:
														myws.cell(row=qd_row, column=j, value=x['value']-x['value'] % (1000 if x['value']//10000 != 0 else 100))
													assign_style(myws.cell(row=qd_row, column=j), font=arial_font, alignment=Alignment(horizontal='left', vertical='top'), border=border)
										elif x['sheet'] == u'Perf Per Watt' and x['name'] == test_row[l]:
											j = curr_col+4+len(test_row)+l
											for c in a:
												if x['test'] == c:
													if not table_iter:
														myws.cell(row=qd_row, column=j, value=x['value'])
													else:
														myws.cell(row=qd_row, column=j, value=x['value']+(1000-x['value'] % 1000 if x['value']//10000 != 0 else 100-x['value'] % 100))
													assign_style(myws.cell(row=qd_row, column=j), font=arial_font, alignment=Alignment(horizontal='left', vertical='top'), border=border)
													temp = 4 + len(test_row) + l + 1
								qd_row += temp_qd.count(e)
							rw_size += temp_rw.count(z)
						size_row += temp_size.count(y)


					if table_iter == 0:
						row = myws.max_row+2 if temp != 0 else 2
					if len([x['test'] for x in row_list if 'Seq' in x['test']]) > 0:
						pattern = PatternFill(start_color='f0f0f0', end_color='f0f0f0', patternType='solid')
						myws.cell(row=row, column=curr_col+1).border = border
						assign_style(myws.cell(row=row, column=curr_col+2), border=border, fill=pattern)
						for x in range(0, 2 * len(test_row)):
							if x < len(test_row):
								myws.cell(row=row, column=curr_col + 3 + x, value=test_row[x].replace('_', ' ') + u' (MiB/s)')
								a = curr_col + 3 + x
							else:
								myws.cell(row=row, column=curr_col + 3 + len(test_row) + x % len(test_row),
										value=test_row[x - len(test_row)].replace('_', ' ') + u' (mW)')
								a = curr_col + 3 + len(test_row) + x % len(test_row)
							assign_style(myws.cell(row=row, column=a), font=bold_font, alignment=alignment, fill=pattern, border=border)

						size_row = row+1
						temp_rw = [x['rw'] for x in row_list if u'Seq' in x['rw']]
						rw = sorted(set(temp_rw), key=lambda x: ('70' in x, 'Write' in x, 'Read' in x))
						for z in rw:
							band_row = size_row
							if z == 'SeqRead':
								value = 'Sequential Read'
							elif z == 'SeqWrite':
								value = 'Sequential Write'
							else:
								value = '70% Read 30% Write'
							temp_size = [x['size'] for x in row_list if u'Seq' in x['rw'] and z == x['rw']]
							size = sorted(set(temp_size), key=lambda x: (x[-1].lower(), len(x), int(x[0])))
							for y in size:
								temp_qd = [x['test'][x['test'].find(u'Qd'):] for x in row_list if u'Seq' in x['rw'] and y == x['size'] and z == x['rw']]
								qd = sorted(set(temp_qd), key=lambda x: (len(x[:x.rfind('_')]), x[:x.rfind('_')], x[x.rfind(u'_')+1]))
								for e in qd:
									myws.cell(row=size_row, column=curr_col + 2, value='{} {} {} ({}W)'.format(y.upper(), e[:e.rfind(u'_')].upper(), value, e[e.rfind(u'_')+1:]))
									assign_style(myws.cell(row=size_row, column=curr_col + 2), font=arial_font, alignment=alignment, border=border)

									a = [x['test'] for x in row_list if x['test'].startswith(y + z + e)]
									for l in range(len(test_row)):
										for x in di:
											if x['sheet'] == u'MiBps_Results' and x['name'] == test_row[l]:
												j = curr_col + 3 + l
												for c in a:
													if x['test'] == c:
														if not table_iter:
															myws.cell(row=size_row, column=j, value=x['value'])
														else:
															myws.cell(row=size_row, column=j,
																	value=x['value'] - x['value'] % (
																	1000 if x['value'] // 10000 != 0 else 100))
														assign_style(myws.cell(row=size_row, column=j), font=arial_font, alignment=Alignment(horizontal='left', vertical='top'), border=border, fill=PatternFill(start_color='e0f0ff', end_color='e0f0ff', patternType='solid'))
											elif x['sheet'] == u'Perf Per Watt' and x['name'] == test_row[l]:
												j = curr_col + 3 + len(test_row) + l
												for c in a:
													if x['test'] == c:
														if not table_iter:
															myws.cell(row=size_row, column=j, value=x['value'])
														else:
															myws.cell(row=size_row, column=j, value=x['value']+(1000-x['value'] % 1000 if x['value'] // 10000 != 0 else 100-x['value'] % 100))
														assign_style(myws.cell(row=size_row, column=j), font=arial_font, border=border, fill=PatternFill(start_color='e0f0ff', end_color='e0f0ff', patternType='solid'), alignment=Alignment(horizontal='left', vertical='top'))
														if temp == 0:
															temp = 3 + len(test_row) + l + 1
									size_row += 1
							myws.cell(row=band_row, column=curr_col+1, value='{} Bandwith'.format(value[value.find(' ')+1:]))
							myws.merge_cells(start_row=band_row, start_column=curr_col + 1, end_row=size_row-1, end_column=curr_col + 1)
							if band_row != size_row-1:
								style_range(myws, '{0}{1}:{0}{2}'.format(
												hex(65 + curr_col)[2:].decode('hex'), band_row, size_row-1),
												font=arial_font, alignment=alignment, border=border)
							else:
								assign_style(myws.cell(row=band_row, column=curr_col + 1), font=arial_font, alignment=alignment, border=border)

					myws.cell(row=1, column=curr_col+1, value='NEW TABLE' if table_iter else 'RAW TABLE')
					myws.merge_cells(start_row=1, start_column=curr_col+1, end_row=1, end_column=curr_col+temp-1)
					style_range(myws, '{0}{1}:{0}{2}'.format(
										hex(65+curr_col)[2:].decode('hex'), 1, 1),
										font=Font(name='Arial', size=11, color='ffffff'), alignment=alignment,
										fill=PatternFill(start_color='000000', end_color='000000', patternType='solid'))
					curr_col += temp

			workbook.save(excel)
		elif user_input in EXIT_COMMANDS:
			print "Program Exiting......."
			break
		else:
			print "INVALID INPUT. TRY AGAIN."
			continue

if __name__ == "__main__":
	if len(sys.argv) == 2:
		main(sys.argv[1])
	else:
		main()
